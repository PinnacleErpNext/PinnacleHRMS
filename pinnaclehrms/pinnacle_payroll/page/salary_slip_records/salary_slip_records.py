import frappe
import calendar
import base64
import json

from datetime import date, timedelta
from io import BytesIO

from frappe import _
from frappe.utils import get_datetime
from frappe.utils.xlsxutils import make_xlsx

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from frappe.utils import flt


@frappe.whitelist()
def getSalarySlipRecords(company, year, month, employee=None):
    year = int(year)
    month = int(month)

    start_date = date(year, month, 1)
    last_day = calendar.monthrange(year, month)[1]
    end_date = date(year, month, last_day)

    filters = {
        "company": company,
        "start_date": [">=", start_date],
        "end_date": ["<=", end_date],
    }

    if employee:
        filters["employee"] = employee

    salary_slips = frappe.get_all("Salary Slip", filters=filters, pluck="name")

    result = []

    for slip_name in salary_slips:
        pay_slip = frappe.get_doc("Salary Slip", slip_name)

        # Salary Breakup
        salary_info = {}
        for sal in pay_slip.salary_breakup:
            salary_info[sal.particulars] = {
                "days": sal.days,
                "amount": sal.amount,
            }

        # Earnings
        basic_salary = 0
        other_earnings_total = 0
        other_earnings_info = []

        for e in pay_slip.earnings:
            if e.salary_component == "Basic":
                basic_salary += e.amount
            else:
                other_earnings_total += e.amount
                other_earnings_info.append(
                    {"component": e.salary_component, "amount": e.amount}
                )

        # Employee info
        emp = (
            frappe.db.get_value(
                "Employee",
                pay_slip.employee,
                ["company_email", "pan_number", "date_of_joining"],
                as_dict=True,
            )
            or {}
        )

        pay_slip_dict = {
            "pay_slip_name": pay_slip.name,
            "status": pay_slip.status,
            "year": year,
            "month": month,
            "employee": pay_slip.employee,
            "employee_name": pay_slip.employee_name,
            "company": pay_slip.company,
            "designation": pay_slip.designation,
            "department": pay_slip.department,
            "email": emp.get("company_email"),
            "standard_working_days": pay_slip.total_working_days,
            "pan_number": emp.get("pan_number"),
            "date_of_joining": emp.get("date_of_joining"),
            "basic_salary": basic_salary,
            "per_day_salary": (
                basic_salary / pay_slip.total_working_days
                if pay_slip.total_working_days
                else 0
            ),
            "actual_working_days": pay_slip.payment_days,
            "absent": pay_slip.absent_days,
            "total": pay_slip.gross_pay,
            "net_payable_amount": pay_slip.net_pay,
            "salary_info": salary_info,
            "other_earnings": other_earnings_info,
            "other_earnings_total": other_earnings_total,
        }

        print(pay_slip_dict)
        result.append(pay_slip_dict)

    return result


@frappe.whitelist()
def download_pay_slip_report(year=None, month=None, encodedCompany=None):
    company = base64.b64decode(encodedCompany).decode("utf-8")

    curr_user = frappe.session.user
    allowed_roles = ["System Manager"]
    user_roles = frappe.get_roles(curr_user)

    if (
        any(role in user_roles for role in allowed_roles)
        and curr_user != "Administrator"
    ):
        frappe.local.response["type"] = "redirect"
        frappe.local.response["location"] = "/app/home"
        return

    if not year or not month:
        frappe.throw(_("Year and Month are required"))

    records = getSalarySlipRecords(company=company, year=year, month=month)

    if not records:
        frappe.throw(_("No data found for the given filters."))

    static_columns = [
        "Pay Slip Name",
        "Year",
        "Month",
        "Employee ID",
        "Employee Name",
        "Company",
        "Designation",
        "Department",
        "Personal Email",
        "Standard Working Days",
        "PAN Number",
        "Date of Joining",
        "Basic Salary",
        "Per Day Salary",
        "Actual Working Days",
        "Absent",
        "Total",
        "Net Payable Amount",
    ]

    salary_info_keys = set()
    other_earning_keys = set()

    for r in records:
        for key in r.get("salary_info", {}).keys():
            if key:
                salary_info_keys.add(key)

        for earning in r.get("other_earnings", []):
            if earning.get("component"):
                other_earning_keys.add(earning.get("component"))

    salary_info_keys = sorted(salary_info_keys)
    other_earning_keys = sorted(other_earning_keys)

    header_row_1 = (
        static_columns
        + ["Salary Info"] * (len(salary_info_keys) * 2)
        + ["Other Earnings"] * len(other_earning_keys)
    )

    header_row_2 = [""] * len(static_columns)

    for key in salary_info_keys:
        header_row_2 += [f"{key} - Days", f"{key} - Amount"]

    for key in other_earning_keys:
        header_row_2.append(f"{key} - Amount")

    data_rows = []

    for r in records:
        row = [
            r.get("pay_slip_name"),
            r.get("year"),
            r.get("month"),
            r.get("employee"),
            r.get("employee_name"),
            r.get("company"),
            r.get("designation"),
            r.get("department"),
            r.get("email"),
            r.get("standard_working_days"),
            r.get("pan_number"),
            r.get("date_of_joining"),
            r.get("basic_salary"),
            r.get("per_day_salary"),
            r.get("actual_working_days"),
            r.get("absent"),
            r.get("total"),
            r.get("net_payable_amount"),
        ]

        salary_info = r.get("salary_info", {})
        for key in salary_info_keys:
            info = salary_info.get(key, {})
            row.append(info.get("days", 0))
            row.append(info.get("amount", 0))

        earnings_map = {
            e.get("component"): e.get("amount")
            for e in r.get("other_earnings", [])
            if e.get("component")
        }

        for key in other_earning_keys:
            row.append(earnings_map.get(key, 0))

        data_rows.append(row)

    xlsx_data = make_xlsx(
        data=[header_row_1, header_row_2] + data_rows,
        sheet_name="Pay Slip Report",
    )

    filename = f"{company.replace(' ', '_')}_Pay_Slip_Report_{calendar.month_name[int(month)]}_{year}.xlsx"

    frappe.response.filename = filename
    frappe.response.filecontent = xlsx_data.getvalue()
    frappe.response.type = "binary"


@frappe.whitelist()
@frappe.whitelist()
def download_idfc_blkpay(year=None, month=None, encodedCompany=None):
    import base64
    import calendar
    from datetime import date
    from io import BytesIO

    import frappe
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    # --- Decode and validate company ---
    company = base64.b64decode(encodedCompany).decode("utf-8")
    company_abbr = frappe.db.get_value("Company", company, "abbr")

    # --- User access check ---
    curr_user = frappe.session.user
    allowed_roles = ["All", "HR User", "HR Manager", "System Manager"]
    user_roles = frappe.get_roles(curr_user)

    if not (
        any(role in user_roles for role in allowed_roles)
        or curr_user == "Administrator"
    ):
        frappe.local.response["type"] = "redirect"
        frappe.local.response["location"] = "/app/home"
        return

    # --- Validate input month & year ---
    if not month or not year:
        frappe.throw("Please specify a month and year")

    try:
        m = int(month)
        year = int(year)

        if not (1 <= m <= 12):
            frappe.throw("Month must be between 1 and 12")

    except ValueError:
        frappe.throw("Invalid month format")

    last_day = calendar.monthrange(year, m)[1]
    last_date = date(year, m, last_day)

    formatted_date_for_filename = last_date.strftime("%d%m%Y")

    # ---------------------------------------------------
    # FETCH DATA FROM SALARY SLIP INSTEAD OF PAY SLIPS
    # ---------------------------------------------------
    query = """
        SELECT
            emp.ifsc_code AS ifsc,
            emp.bank_ac_no AS beneficiary_account_no,
            ss.employee_name AS beneficiary_name,
            ss.company,
            ss.net_pay AS amount
        FROM `tabSalary Slip` ss
        INNER JOIN `tabEmployee` emp
            ON emp.name = ss.employee
        WHERE
            yMONTH(ss.end_date) = %(month)s
            AND YEAR(ss.end_date) = %(year)s
            AND ss.company = %(company)s
            AND IFNULL(ss.net_pay, 0) > 0
    """

    params = {
        "month": m,
        "year": year,
        "company": company,
    }

    data = frappe.db.sql(query, params, as_dict=True)

    if not data:
        frappe.throw("No submitted Salary Slips found for selected month and company")

    # --- Debit account map ---
    company_debit_map = {
        "Opticodes Technologies Private Limited": "10238672140",
        "Pinnacle Finserv Advisors Pvt. Ltd.": "10237782223",
    }

    today = date.today()

    # --- Column configuration ---
    columns = [
        {"header": "Beneficiary Name", "width": 30},
        {"header": "Beneficiary Account No", "width": 25},
        {"header": "IFSC", "width": 20},
        {"header": "Transaction Type", "width": 20},
        {"header": "Debit Account Number", "width": 30},
        {"header": "Transaction Date", "width": 20},
        {"header": "Amount (₹)", "width": 15},
        {"header": "Currency", "width": 10},
    ]

    # --- Prepare row data ---
    rows = []

    for r in data:
        debit_account = company_debit_map.get(r.company, "")

        rows.append(
            [
                r.beneficiary_name or "",
                str(r.beneficiary_account_no or ""),
                r.ifsc or "",
                "NEFT",
                debit_account,
                today.strftime("%d/%m/%Y"),
                flt(r.amount, 2),
                "INR",
            ]
        )

    # --- Build Excel file ---
    wb = Workbook()
    ws = wb.active
    ws.title = "IDFC BLKPAY"

    # --- Header Row ---
    for col_num, col in enumerate(columns, 1):
        col_letter = ws.cell(row=1, column=col_num).column_letter

        ws.merge_cells(f"{col_letter}1:{col_letter}2")

        cell = ws.cell(row=1, column=col_num, value=col["header"])

        ws.column_dimensions[col_letter].width = col["width"]

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        cell.font = Font(bold=True)

    # --- Write data rows ---
    start_row = 3

    for row_idx, row_data in enumerate(rows, start_row):

        for col_idx, value in enumerate(row_data, 1):

            header = columns[col_idx - 1]["header"]

            cell = ws.cell(
                row=row_idx,
                column=col_idx,
                value=value,
            )

            # Keep account numbers as text
            if header in ["Beneficiary Account No", "IFSC"]:
                cell.number_format = "@"

            elif header == "Amount (₹)":
                cell.number_format = "0.00"
                cell.alignment = Alignment(horizontal="right")

            else:
                cell.alignment = Alignment(horizontal="left")

    # --- Save workbook to memory ---
    output = BytesIO()

    wb.save(output)

    output.seek(0)

    # --- Set response ---
    frappe.response.filename = f"{company_abbr}{formatted_date_for_filename}.xlsx"

    frappe.response.filecontent = output.getvalue()
    frappe.response.type = "binary"
