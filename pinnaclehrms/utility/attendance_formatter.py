# -*- coding: utf-8 -*-
import io
import json
from datetime import datetime, time
from collections import defaultdict

import frappe
from frappe.utils.file_manager import save_file
from werkzeug.wrappers import Response

import openpyxl
from openpyxl import load_workbook, Workbook


# ============================================================
# ------------------------ Helpers ---------------------------
# ============================================================

MISSING_TIME_STRINGS = {
    "",
    " ",
    "None",
    "none",
    "NULL",
    "null",
    "NaT",
    "00:00",
    "00:00:00",
    "0:0",
    "0:00",
    "00:0",
}


def is_missing_time(val) -> bool:
    """Return True if a time value is blank/zero-like."""
    if val is None:
        return True
    if isinstance(val, time):
        return False
    s = str(val).strip()
    return s in MISSING_TIME_STRINGS


def merge_header_cells(header_cells):
    """Merge split header like ['Attendance','Device','Id'] -> 'Attendance Device Id'"""
    merged, temp = [], []
    for val in header_cells:
        if val is None:
            continue
        text = str(val).strip()
        if text.lower() in [
            "attendance",
            "device",
            "id",
            "employee",
            "name",
            "date",
            "shift",
            "in",
            "out",
            "time",
        ]:
            temp.append(text)
            if text.lower() in ["id", "name", "date", "shift", "time"]:
                merged.append(" ".join(temp))
                temp = []
        else:
            merged.append(text)
    return merged


def parse_date_safe(date_val):
    """Try multiple formats into datetime (retain date)."""
    if isinstance(date_val, datetime):
        return date_val
    for fmt in ("%Y-%m-%d", "%d-%b-%Y", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(str(date_val), fmt)
        except Exception:
            continue
    # last resort: pandas-like parser
    try:
        return datetime.fromisoformat(str(date_val))
    except Exception:
        return None


def parse_time_safe(value):
    """Parse time/datetime strings -> time()."""
    if isinstance(value, datetime):
        return value.time()
    if isinstance(value, time):
        return value
    if value is None:
        return None
    s = str(value).strip()
    if s in MISSING_TIME_STRINGS:
        return None

    formats = [
        "%H:%M:%S",
        "%H:%M",
        "%H:%M:%S.%f",
        "%I:%M:%S %p",
        "%I:%M %p",
        "%I:%M:%S.%f %p",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d %H:%M:%S.%f",
        "%Y-%m-%d %I:%M:%S %p",
        "%Y-%m-%d %I:%M %p",
        "%Y-%m-%d %I:%M:%S.%f %p",
        "%Y/%m/%d %H:%M:%S",
        "%Y/%m/%d %H:%M",
        "%Y/%m/%d %H:%M:%S.%f",
        "%Y/%m/%d %I:%M:%S %p",
        "%Y/%m/%d %I:%M %p",
        "%Y/%m/%d %I:%M:%S.%f %p",
        "%d/%m/%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d-%m-%Y %H:%M:%S",
        "%d-%m-%Y %H:%M",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(s, fmt).time()
        except ValueError:
            continue
    # last resort
    try:
        dt = datetime.fromisoformat(s)
        return dt.time()
    except Exception:
        return None


def format_time_string(t):
    if isinstance(t, datetime):
        return t.strftime("%H:%M:%S")
    elif isinstance(t, time):
        return t.strftime("%H:%M:%S")
    elif t:
        return str(t).strip()
    return None


def get_employee(company=None):
    """Return {employee_id: employee_name}"""
    filters = {"status": "Active"}
    if company:
        filters["company"] = company
    employees = frappe.get_all(
        "Employee", filters=filters, fields=["name", "employee_name"]
    )
    return {emp.name: emp.employee_name for emp in employees}


def map_device_to_employee(device, device_id):
    """Map Attendance Device+ID to Employee via 'Attendance Device ID Allotment' child table."""
    if not device or not device_id:
        return None
    return frappe.db.get_value(
        "Attendance Device ID Allotment",
        {"device": device, "device_id": str(device_id)},
        "parent",
    )


def get_app_attendance(employee_list, payrollFrom, payrollTo):
    """
    Fetch attendance from Employee Checkin (App source) for a list of employees in date range.
    Returns { employee_id: [ {employee, employee_name, shift, attendance_date, in_time, out_time} ] }
    """
    if not employee_list:
        return {}

    query = """
        SELECT
            employee,
            employee_name,
            shift,
            DATE(`time`) AS attendance_date,
            MIN(CASE WHEN log_type = 'IN'  THEN `time` END)  AS in_time,
            MAX(CASE WHEN log_type = 'OUT' THEN `time` END)  AS out_time
        FROM `tabEmployee Checkin`
        WHERE employee IN %(employee_list)s
          AND DATE(`time`) BETWEEN %(from_date)s AND %(to_date)s
        GROUP BY employee, DATE(`time`)
        ORDER BY employee, attendance_date
    """

    rows = frappe.db.sql(
        query,
        {
            "employee_list": tuple(employee_list),
            "from_date": payrollFrom,
            "to_date": payrollTo,
        },
        as_dict=True,
    )

    attendance = defaultdict(list)
    for r in rows:
        attendance[r["employee"]].append(
            {
                "device_id": "",
                "device": "App",
                "employee_name": r.get("employee_name"),
                "employee_id": r.get("employee"),
                "attendance_date": r.get("attendance_date"),
                "shift": r.get("shift") or "Regular",
                "in_time": r.get("in_time"),
                "out_time": r.get("out_time"),
            }
        )
    return dict(attendance)


def convert_app_attendance_to_records(app_attendance):
    """Flatten App attendance map to list of device-like rows."""
    out = []
    for emp_id, records in app_attendance.items():
        for rec in records:
            out.append(
                {
                    "device_id": "",  # app has no device id
                    "device": "App",
                    "employee_name": rec.get("employee_name"),
                    "employee_id": emp_id,
                    "attendance_date": rec.get("attendance_date"),
                    "shift": rec.get("shift") or "Regular",
                    "in_time": rec.get("in_time"),
                    "out_time": rec.get("out_time"),
                }
            )
    return out


# ============================================================
# ----------------------- Processors -------------------------
# ============================================================


def process_pinnacle(file):
    """Zicom Regal sheet 'Att.log report' -> list of dict rows."""
    file_stream = file.stream.read()
    wb = load_workbook(filename=io.BytesIO(file_stream), data_only=True)
    if "Att.log report" not in wb.sheetnames:
        frappe.throw("Sheet 'Att.log report' not found.")
    ws = wb["Att.log report"]

    raw_period = ws["C3"].value
    if not raw_period or "~" not in str(raw_period):
        frappe.throw("Invalid period in C3 for Pinnacle file.")
    start_date_str = str(raw_period).split("~")[0].strip()
    start_date = parse_date_safe(start_date_str)
    if not start_date:
        frappe.throw("Unable to parse start date in Pinnacle file.")

    formatted_period = start_date.strftime("%b-%Y")
    dates = [cell.value for cell in ws[4] if isinstance(cell.value, int)]

    records = []
    row = 5
    while ws.cell(row=row, column=1).value:
        if ws.cell(row=row, column=1).value == "ID:":
            device_id = str(ws.cell(row=row, column=3).value or "").strip()
            emp_name = str(ws.cell(row=row, column=11).value or "").strip()
            time_log_row = row + 1

            for col_index, day in enumerate(dates, start=1):
                time_log = ws.cell(row=time_log_row, column=col_index).value
                if isinstance(time_log, str):
                    time_log = time_log.strip()
                    # Often format looks like 'HH:MM-HH:MM' or 'HH:MM   HH:MM'
                    in_time = time_log[:5]
                    out_time = time_log[-5:] if len(time_log) >= 10 else ""

                    date_obj = datetime.strptime(
                        f"{int(day):02d}-{formatted_period}", "%d-%b-%Y"
                    )
                    records.append(
                        {
                            "device_id": device_id,
                            "device": "Zicom Regal",
                            "employee_name": emp_name,
                            "attendance_date": date_obj.strftime("%d-%b-%Y"),
                            "shift": "Regular",
                            "in_time": in_time,
                            "out_time": out_time,
                        }
                    )
            row += 2
        else:
            row += 1
    return records


def process_Opticode_final(file):
    """ESSL Westcott: sheet 'Final' -> list of device rows."""
    file_stream = file.stream.read()
    wb = load_workbook(filename=io.BytesIO(file_stream), data_only=True)
    records, seen = [], set()

    if "Final" not in wb.sheetnames:
        frappe.throw("Sheet 'Final' not found in the workbook.")
    sheet = wb["Final"]
    if sheet.max_row < 2:
        return []

    header_row = [
        cell.value.strip() if isinstance(cell.value, str) else cell.value
        for cell in sheet[1]
    ]
    col_index = {header: idx for idx, header in enumerate(header_row)}

    required_fields = ["ID", "G", "Date", "In Time", "Out Time"]
    for field in required_fields:
        if field not in col_index:
            frappe.throw(f"Missing required column: {field}")

    for row in sheet.iter_rows(min_row=2, values_only=True):
        try:
            device_id = row[col_index["ID"]]
            emp_name = row[col_index["G"]]
            date_val = row[col_index["Date"]]
            in_time = row[col_index["In Time"]]
            out_time = row[col_index["Out Time"]]

            if not (device_id and emp_name and date_val):
                continue

            # Normalize date
            if isinstance(date_val, datetime):
                date_key = date_val.date()
                date_str = date_val.strftime("%d-%b-%Y")
            else:
                date_obj = parse_date_safe(date_val)
                if not date_obj:
                    continue
                date_key = date_obj.date()
                date_str = date_obj.strftime("%d-%b-%Y")

            in_time_str = format_time_string(parse_time_safe(in_time)) or ""
            out_time_str = format_time_string(parse_time_safe(out_time)) or ""

            unique_key = (str(device_id), date_key, in_time_str, out_time_str)
            if unique_key in seen:
                continue
            seen.add(unique_key)

            records.append(
                {
                    "device_id": str(device_id),
                    "device": "ESSL Westcott",
                    "employee_name": str(emp_name).strip(),
                    "attendance_date": date_str,
                    "shift": "Regular",
                    "in_time": in_time_str,
                    "out_time": out_time_str,
                }
            )
        except Exception as e:
            frappe.log_error(
                f"Error processing Opticode row: {e}", "Opticode Formatter"
            )
    return records


def process_mantra(file):
    """Mantra generic dump (first row header)."""
    file_stream = file.stream.read()
    wb = load_workbook(filename=io.BytesIO(file_stream), data_only=True)
    sheet = wb.active

    raw_header = [c.value for c in sheet[1]]
    header = merge_header_cells(raw_header)
    header = [h.lower().strip() for h in header]
    col_idx = {h: i for i, h in enumerate(header)}

    required = [
        "attendance device id",
        "attendance device",
        "employee name",
        "attendance date",
        "in time",
        "out time",
    ]
    for col in required:
        if col not in col_idx:
            frappe.throw(f"Missing required column: {col}")

    records, seen = [], set()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        vals = [v for v in row if v not in (None, "", " ")]
        if not vals:
            continue

        device_id = str(row[col_idx["attendance device id"]] or "").strip()
        device = str(row[col_idx["attendance device"]] or "").strip()
        emp_name = str(row[col_idx["employee name"]] or "").strip()
        date_val = row[col_idx["attendance date"]]
        in_val = row[col_idx["in time"]]
        out_val = row[col_idx["out time"]]

        if not (device_id and emp_name and date_val):
            continue

        date_obj = parse_date_safe(date_val)
        if not date_obj:
            continue
        attendance_date = date_obj.strftime("%d-%b-%Y")

        in_time = format_time_string(parse_time_safe(in_val)) or ""
        out_time = format_time_string(parse_time_safe(out_val)) or ""

        key = (device_id, emp_name, attendance_date, in_time, out_time)
        if key in seen:
            continue
        seen.add(key)

        records.append(
            {
                "device_id": device_id,
                "device": device,
                "employee_name": emp_name,
                "department": "",
                "attendance_date": attendance_date,
                "shift": "Regular",
                "in_time": in_time,
                "out_time": out_time,
            }
        )
    return records


def process_other(file):
    """Custom 'Other' format with headers: Employee, Employee Name, Attendance Date, In Time, Out Time."""
    file_stream = file.stream.read()
    wb = load_workbook(filename=io.BytesIO(file_stream), data_only=True)
    sheet = wb.active
    records, seen = [], set()

    if sheet.max_row < 2:
        return []

    header_row = [
        cell.value.strip() if isinstance(cell.value, str) else cell.value
        for cell in sheet[1]
    ]
    col_index = {header: idx for idx, header in enumerate(header_row)}

    required_fields = [
        "Employee",
        "Employee Name",
        "Attendance Date",
        "In Time",
        "Out Time",
    ]
    for field in required_fields:
        if field not in col_index:
            frappe.throw(f"Missing required column: {field}")

    for row in sheet.iter_rows(min_row=2, values_only=True):
        try:
            employee = row[col_index["Employee"]]
            emp_name = row[col_index["Employee Name"]]
            date_val = row[col_index["Attendance Date"]]
            in_time = row[col_index["In Time"]]
            out_time = row[col_index["Out Time"]]

            if not (employee and emp_name and date_val):
                continue

            date_obj = parse_date_safe(date_val)
            if not date_obj:
                continue
            date_key = date_obj.date()
            date_str = date_obj.strftime("%d-%b-%Y")

            in_time_str = format_time_string(parse_time_safe(in_time)) or ""
            out_time_str = format_time_string(parse_time_safe(out_time)) or ""

            unique_key = (str(employee), date_key, in_time_str, out_time_str)
            if unique_key in seen:
                continue
            seen.add(unique_key)

            records.append(
                {
                    "employee_id": str(employee),
                    "employee_name": str(emp_name).strip(),
                    "attendance_date": date_str,
                    "shift": "Regular",
                    "in_time": in_time_str,
                    "out_time": out_time_str,
                }
            )
        except Exception as e:
            frappe.log_error(f"Error processing Other row: {e}", "Other Formatter")
    return records


# ============================================================
# --------------------- Core Generator -----------------------
# ============================================================


def _normalize_record(record):
    """
    Convert any record (device/app/other) into a normalized tuple:
    (employee_id, employee_name, date(date), in_time(time|None), out_time(time|None), src_in, src_out)
    """
    device = record.get("device")
    device_id = record.get("device_id")
    employee_name = record.get("employee_name") or ""
    raw_emp_id = record.get("employee_id")

    # Resolve employee id
    if device == "App" or not device:
        emp_id = raw_emp_id
    else:
        emp_id = raw_emp_id or map_device_to_employee(device, device_id)

    if not emp_id:
        return None  # can't map to employee -> skip

    # Parse date & times
    dt = parse_date_safe(record.get("attendance_date"))
    if not dt:
        return None
    d = dt.date()

    in_raw = record.get("in_time")
    out_raw = record.get("out_time")
    in_t = None if is_missing_time(in_raw) else parse_time_safe(in_raw)
    out_t = None if is_missing_time(out_raw) else parse_time_safe(out_raw)

    # Sources (for display/debug)
    src_in = "App" if (in_t and device == "App") else (device or "")
    src_out = "App" if (out_t and device == "App") else (device or "")

    return emp_id, employee_name, d, in_t, out_t, src_in, src_out


@frappe.whitelist()
def generate_final_sheet(attendance_data=None, use_clubbed_punch_logic=True):
    """
    Generate final attendance sheet.

    LOGIC 1 (default / new):
      - Group by employee + date
      - Club ALL punches (device + app)
      - Sort punches
      - First punch  -> IN
      - Last punch   -> OUT

    LOGIC 2 (old):
      - Earliest IN, Latest OUT (direction-based)

    EXTRA FIX:
      - If IN == OUT → fetch App punches, re-mix & recalc
    """

    def get_app_punches(emp_id, d):
        punches = []
        rows = frappe.get_all(
            "Employee Checkin",
            filters={
                "employee": emp_id,
                "time": ["between", [f"{d} 00:00:00", f"{d} 23:59:59"]],
            },
            fields=["time"],
            order_by="time asc",
        )

        for r in rows:
            punches.append(
                {
                    "time": r.time.time(),
                    "src": "App",
                }
            )
        return punches

    attendance_by_emp = defaultdict(list)
    emp_name_cache = {}

    # --------------------------------------------------
    # 1. Normalize & collect raw rows
    # --------------------------------------------------
    for data in attendance_data or []:
        try:
            norm = _normalize_record(data)
            if not norm:
                continue

            emp_id, emp_name, d, in_t, out_t, src_in, src_out = norm

            if emp_id not in emp_name_cache:
                emp_name_cache[emp_id] = emp_name or frappe.db.get_value(
                    "Employee", emp_id, "employee_name"
                )

            attendance_by_emp[emp_id].append(
                {
                    "date": d,
                    "shift": (data.get("shift") or "Regular").strip(),
                    "in_time": in_t,
                    "out_time": out_t,
                    "src_in": src_in or data.get("device", "") or "",
                    "src_out": src_out or data.get("device", "") or "",
                }
            )

        except Exception:
            frappe.log_error(
                frappe.get_traceback(),
                f"Error processing entry: {data}",
            )

    # --------------------------------------------------
    # 2. Generate final data
    # --------------------------------------------------
    final_data = {}

    for emp_id, rows in attendance_by_emp.items():
        emp_name = emp_name_cache.get(emp_id)
        by_date = defaultdict(list)

        for r in rows:
            by_date[r["date"]].append(r)

        result_rows = []

        for d in sorted(by_date.keys()):
            day_rows = by_date[d]

            punches = []

            # ===============================
            # 🔥 NEW LOGIC (Club punches)
            # ===============================
            if use_clubbed_punch_logic:
                for r in day_rows:
                    if r["in_time"]:
                        punches.append({"time": r["in_time"], "src": r["src_in"]})
                    if r["out_time"]:
                        punches.append({"time": r["out_time"], "src": r["src_out"]})

            # ===============================
            # 🟡 OLD LOGIC (Direction-based)
            # ===============================
            else:
                in_time = None
                out_time = None
                src_in = ""
                src_out = ""

                for r in day_rows:
                    if r["in_time"] and (not in_time or r["in_time"] < in_time):
                        in_time = r["in_time"]
                        src_in = r["src_in"]

                    if r["out_time"] and (not out_time or r["out_time"] > out_time):
                        out_time = r["out_time"]
                        src_out = r["src_out"]

                if in_time or out_time:
                    punches.append({"time": in_time, "src": src_in})
                    punches.append({"time": out_time, "src": src_out})

            # --------------------------------------------------
            # 3. Fix SAME IN & OUT → Mix App punches
            # --------------------------------------------------
            punches.extend(get_app_punches(emp_id, d))

            punches = [p for p in punches if p.get("time")]

            punches.sort(key=lambda x: x["time"])

            in_time = punches[0]["time"] if punches else None
            out_time = punches[-1]["time"] if punches else None
            src_in = punches[0]["src"] if punches else ""
            src_out = punches[-1]["src"] if punches else ""

            # --------------------------------------------------
            # 4. Final Row
            # --------------------------------------------------
            result_rows.append(
                {
                    "employee": emp_id,
                    "employee_name": emp_name,
                    "attendance_date": d,
                    "shift": "Regular",
                    "custom_log_in_from": src_in,
                    "in_time": format_time_string(in_time) or "",
                    "custom_log_out_from": src_out,
                    "out_time": format_time_string(out_time) or "",
                }
            )

        final_data[emp_id] = result_rows

    return {
        "message": "✅ Attendance extracted successfully",
        "logic_used": (
            "clubbed_punch_logic"
            if use_clubbed_punch_logic
            else "direction_based_logic"
        ),
        "total_employees": len(final_data),
        "total_records": sum(len(v) for v in final_data.values()),
        "data": final_data,
    }


# ============================================================
# --------------------- Load Raw Data ------------------------
# ============================================================


@frappe.whitelist()
def load_raw_attendance_data():
    pinnacle_file = frappe.request.files.get("pinnacle_file")
    opticode_file = frappe.request.files.get("opticode_file")
    mantra_file = frappe.request.files.get("mantra_file")
    other_file = frappe.request.files.get("other_file")

    payrollFrom = frappe.form_dict.get("from_date")
    payrollTo = frappe.form_dict.get("to_date")

    if not payrollFrom or not payrollTo:
        return Response("❌ Payroll Period is required", status=400)

    employeeList = get_employee()
    if not employeeList:
        return Response("❌ No active employees found", status=404)

    # App attendance pull for ALL employees in period
    app_attendance = get_app_attendance(
        list(employeeList.keys()), payrollFrom, payrollTo
    )
    app_records = convert_app_attendance_to_records(app_attendance)

    # Safe defaults
    pinnacleAttendance, opticodeAttendance, mantraAttendance, otherFile = [], [], [], []

    try:
        if pinnacle_file:
            pinnacleAttendance = process_pinnacle(pinnacle_file)
    except Exception:
        frappe.log_error(frappe.get_traceback(), "Pinnacle Processing Error")

    try:
        if opticode_file:
            opticodeAttendance = process_Opticode_final(opticode_file)
    except Exception:
        frappe.log_error(frappe.get_traceback(), "Opticode Processing Error")

    try:
        if mantra_file:
            mantraAttendance = process_mantra(mantra_file)
    except Exception:
        frappe.log_error(frappe.get_traceback(), "Mantra Processing Error")

    try:
        if other_file:
            otherFile = process_other(other_file)
    except Exception:
        frappe.log_error(frappe.get_traceback(), "Other Processing Error")

    return {
        "message": "✅ Attendance files processed successfully",
        "status_cd": 200,
        "pinnacle_attendance": pinnacleAttendance,
        "opticode_attendance": opticodeAttendance,
        "mantra_attendance": mantraAttendance,
        "other_attendance": otherFile,
        "app_attendance": app_records,
    }


# ============================================================
# ------------ Preview & Download Endpoints ------------------
# ============================================================


def _coerce_json_arg(payload):
    """Accept dict (already parsed by Frappe) OR JSON string."""
    if payload is None:
        return None
    if isinstance(payload, (dict, list)):
        return payload
    try:
        return json.loads(payload)
    except Exception:
        return None


@frappe.whitelist()
def preview_final_attendance_sheet(raw_data=None):
    """
    Accepts: raw_data = {
      "pinnacle": [...],
      "opticode": [...],
      "mantra": [...],
      "other": [...],
      "app": [...]
    }
    Returns aggregated preview data like generate_final_sheet
    """
    raw_data = _coerce_json_arg(raw_data) or {}
    records = []
    records += raw_data.get("pinnacle", []) or []
    records += raw_data.get("opticode", []) or []
    records += raw_data.get("mantra", []) or []
    records += raw_data.get("other", []) or []
    records += raw_data.get("app", []) or []

    result = generate_final_sheet(records)

    data = result.get("data", {})
    return {
        "message": "Preview generated successfully",
        "data": data,
        "total_employees": len(data),
        "total_records": sum(len(x) for x in data.values()),
    }


@frappe.whitelist()
def download_final_attendance_excel(logs):
    """
    Accepts:
      - Device-style list (with 'device' key)
      - Aggregated dict keyed by employee (final preview/validated)
    Returns Excel file.
    """
    data = _coerce_json_arg(logs)

    wb = Workbook()
    ws = wb.active
    ws.title = "Final Attendance"

    if (
        isinstance(data, list)
        and data
        and isinstance(data[0], dict)
        and "device" in data[0]
    ):
        # Device feed header
        ws.append(
            [
                "Device Id",
                "Device",
                "Employee Name",
                "Attendance Date",
                "Shift",
                "In Time",
                "Out Time",
            ]
        )
        for row in data:
            ws.append(
                [
                    row.get("device_id", ""),
                    row.get("device", ""),
                    row.get("employee_name", ""),
                    (
                        row.get("attendance_date")
                        if isinstance(row.get("attendance_date"), str)
                        else (
                            row.get("attendance_date").strftime("%d-%b-%Y")
                            if row.get("attendance_date")
                            else ""
                        )
                    ),
                    row.get("shift", ""),
                    row.get("in_time", ""),
                    row.get("out_time", ""),
                ]
            )
    else:
        # Aggregated employee dict
        ws.append(
            [
                "Employee",
                "Employee Name",
                "Attendance Date",
                "Shift",
                "Log In From",
                "In Time",
                "Log Out From",
                "Out Time",
            ]
        )
        if isinstance(data, dict):
            for emp_id in sorted(data):
                for row in data[emp_id]:
                    ws.append(
                        [
                            row.get("employee", ""),
                            row.get("employee_name", ""),
                            (
                                row.get("attendance_date").strftime("%d-%b-%Y")
                                if isinstance(row.get("attendance_date"), datetime)
                                else str(row.get("attendance_date") or "")
                            ),
                            row.get("shift", ""),
                            row.get("custom_log_in_from", ""),
                            row.get("in_time", ""),
                            row.get("custom_log_out_from", ""),
                            row.get("out_time", ""),
                        ]
                    )

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Final_Attendance.xlsx"},
    )


# ============================================================
# ------------- Data Import (Validated Records) --------------
# ============================================================


@frappe.whitelist()
def create_data_import_for_attendance(attendance_data=None):
    """
    Creates a Data Import record with an Excel file for Attendance List and starts the import.
    `attendance_data` can be dict keyed by employee id.
    """
    validated_data = _coerce_json_arg(attendance_data) or {}
    if not validated_data:
        frappe.throw("No validated records found.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance List"

    headers = [
        "Employee",
        "Employee Name",
        "Attendance Date",
        "Shift",
        "Log In From",
        "In Time",
        "Log Out From",
        "Out Time",
    ]
    ws.append(headers)

    for emp_id in sorted(validated_data):
        for row in validated_data[emp_id]:
            ws.append(
                [
                    row.get("employee"),
                    row.get("employee_name"),
                    (
                        row.get("attendance_date").strftime("%Y-%m-%d")
                        if isinstance(row.get("attendance_date"), datetime)
                        else str(row.get("attendance_date") or "")
                    ),
                    row.get("shift"),
                    row.get("custom_log_in_from"),
                    row.get("in_time"),
                    row.get("custom_log_out_from"),
                    row.get("out_time"),
                ]
            )

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    data_import = frappe.get_doc(
        {
            "doctype": "Data Import",
            "import_type": "Insert New Records",
            "reference_doctype": "Attendance",
            "submit_after_import": 0,
            "mute_emails": 1,
        }
    )
    data_import.insert(ignore_permissions=True)

    file_doc = save_file(
        "attendance_import.xlsx",
        output.getvalue(),
        "Data Import",
        data_import.name,
        is_private=1,
    )
    data_import.import_file = file_doc.file_url
    data_import.save(ignore_permissions=True)

    frappe.enqueue(
        "frappe.core.doctype.data_import.data_import.start_import",
        data_import=data_import.name,
        import_type="Insert New Records",
    )

    return data_import.name


# ============================================================
# ------------------ Validated Records -----------------------
# ============================================================


@frappe.whitelist()
def validate_attendance_data(attendance_data=None):
    data = _coerce_json_arg(attendance_data)
    if not data:
        return {"validated": {}, "non_validated": []}

    # Normalize input
    if isinstance(data, dict):
        rows = []
        for emp in data:
            rows.extend(data[emp])
    else:
        rows = data

    validated = defaultdict(list)
    non_validated = []
    seen_dates = set()

    # ---------------- RULES ----------------

    def is_empty_or_zero(t):
        if t is None:
            return True
        t = str(t).strip()
        return t == "" or t in ("0", "00", "00:00", "00:00:00")

    def rule_missing_punch(r):
        if is_empty_or_zero(r.get("in_time")) or is_empty_or_zero(r.get("out_time")):
            return False, "Missing IN or OUT time"
        return True, None

    def rule_same_punch(r):
        if r.get("in_time") and r.get("out_time"):
            if str(r["in_time"]).strip() == str(r["out_time"]).strip():
                return False, "IN time and OUT time are same"
        return True, None

    def rule_invalid_time(r):
        try:
            in_raw = r.get("in_time")
            out_raw = r.get("out_time")

            if is_empty_or_zero(in_raw) or is_empty_or_zero(out_raw):
                return True, None

            in_time = parse_time_safe(str(in_raw))
            out_time = parse_time_safe(str(out_raw))

            if in_time is None or out_time is None:
                return False, "Invalid time format"

            if in_time > out_time:
                return False, "IN time is after OUT time"

        except Exception:
            return False, "Invalid time format"

        return True, None

    def rule_duplicate(r):
        key = f"{r.get('employee')}_{r.get('attendance_date')}"
        if key in seen_dates:
            return False, "Duplicate entry"
        seen_dates.add(key)
        return True, None

    # ---------------- PIPELINE ----------------

    validations = [
        rule_missing_punch,
        rule_same_punch,
        rule_invalid_time,
        rule_duplicate,
    ]

    for r in rows:
        errors = []

        for check in validations:
            ok, msg = check(r)
            if not ok:
                errors.append(msg)

        if errors:
            r["errors"] = errors
            non_validated.append(r)
        else:
            validated[r.get("employee")].append(r)

    return {
        "message": "Validation completed",
        "validated": validated,
        "non_validated": non_validated,
        "total_valid": sum(len(v) for v in validated.values()),
        "total_invalid": len(non_validated),
    }
