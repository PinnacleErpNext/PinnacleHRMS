import frappe
import io
import json
from datetime import datetime, time
from collections import defaultdict
from openpyxl import load_workbook, Workbook
from werkzeug.wrappers import Response
from collections import defaultdict
from datetime import time
import openpyxl
from io import BytesIO
from frappe.utils.file_manager import save_file
from collections import defaultdict
from frappe.utils.data import getdate


# --- Helpers ---
def _coerce_json_arg(arg):
    if not arg:
        return None
    if isinstance(arg, str):
        try:
            return json.loads(arg)
        except Exception:
            return None
    return arg


def _get_emp_id(device, device_id):
    if not device or not device_id:
        return None
    return frappe.db.get_value(
        "Attendance Device ID Allotment",
        {"device": device, "device_id": device_id},
        "parent",
    )


def merge_header_cells(header_cells):
    merged = []
    temp = []
    for val in header_cells:
        if val is None:
            continue
        text = str(val).strip()
        # If cell is a word and next cell continues header
        if text.lower() in [
            "attendance",
            "device",
            "id",
            "employee",
            "name",
            "date",
            "shift",
            "in",
            "out",
            "time",
        ]:
            temp.append(text)
            # If the word completes a header pair, join it
            if text.lower() in ["id", "name", "date", "shift", "time"]:
                merged.append(" ".join(temp))
                temp = []
        else:
            merged.append(text)
    return merged


def format_date_safe(date_value):
    """
    Safely format a date value to 'YYYY-MM-DD'.
    Handles datetime objects, strings, or None.
    Returns an empty string if the date is invalid.
    """
    if not date_value:
        return ""

    # If already a datetime or date object
    if isinstance(date_value, (datetime,)):
        return date_value.strftime("%Y-%m-%d")

    # If it's a string, try to parse it
    try:
        parsed_date = datetime.strptime(str(date_value), "%Y-%m-%d")
        return parsed_date.strftime("%Y-%m-%d")
    except ValueError:
        # If parsing fails, just return the original string
        return str(date_value)


def convert_app_attendance_to_records(app_attendance):
    """Convert app_attendance dict into records list matching final sheet format."""
    converted_records = []
    for emp_id, records in app_attendance.items():
        for rec in records:
            converted_records.append(
                {
                    "device_id": "",  # App logs don't have device_id
                    "device": "App",  # Mark source as App
                    "employee_name": rec.get("employee_name"),
                    "employee_id": emp_id,
                    "attendance_date": rec.get("attendance_date"),
                    "shift": rec.get("shift") or "Regular",
                    "in_time": rec.get("in_time"),
                    "out_time": rec.get("out_time"),
                }
            )
    return converted_records


def parse_date_safe(date_val):
    if isinstance(date_val, datetime):
        return date_val
    for fmt in ("%Y-%m-%d", "%d-%b-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(str(date_val), fmt)
        except Exception:
            continue
    return None


from datetime import datetime, time


def parse_time_safe(value):
    """
    Parse various time or datetime string formats safely and return a time object.
    Returns None if parsing fails.
    """
    if isinstance(value, datetime):
        return value.time()

    if isinstance(value, time):
        return value

    if isinstance(value, str):
        value = value.strip()

        # Common formats for time and datetime strings
        formats = [
            # 24-hour time formats
            "%H:%M:%S",
            "%H:%M",
            "%H:%M:%S.%f",
            # 12-hour time formats
            "%I:%M:%S %p",
            "%I:%M %p",
            "%I:%M:%S.%f %p",
            # Date + time (24-hour)
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d %H:%M",
            "%Y-%m-%d %H:%M:%S.%f",
            # Date + time (12-hour)
            "%Y-%m-%d %I:%M:%S %p",
            "%Y-%m-%d %I:%M %p",
            "%Y-%m-%d %I:%M:%S.%f %p",
            # Alternate separators (e.g., slashes)
            "%Y/%m/%d %H:%M:%S",
            "%Y/%m/%d %H:%M",
            "%Y/%m/%d %H:%M:%S.%f",
            "%Y/%m/%d %I:%M:%S %p",
            "%Y/%m/%d %I:%M %p",
            "%Y/%m/%d %I:%M:%S.%f %p",
        ]

        for fmt in formats:
            try:
                return datetime.strptime(value, fmt).time()
            except ValueError:
                continue

    return None


def format_time_string(t):
    if isinstance(t, datetime):
        return t.strftime("%H:%M:%S")
    elif t:
        return str(t).strip()
    return None


def get_employee(company=None):
    if company is None:
        employees = frappe.get_all(
            "Employee",
            filters={"status": "Active"},
            fields=["name", "employee_name"],
        )
        return {emp.name: emp.employee_name for emp in employees}
    employees = frappe.get_all(
        "Employee",
        filters={"company": company, "status": "Active"},
        fields=["name", "employee_name"],
    )
    return {emp.name: emp.employee_name for emp in employees}


from collections import defaultdict


def get_app_attendance(employee_list, payrollFrom, payrollTo):
    """
    Fetch RAW attendance punches for employees within a date range.

    Returns ALL punches with:
    - employee
    - employee_name
    - shift
    - attendance_date
    - time
    - log_type

    Only includes records where skip_auto_attendance = 0
    """

    if not employee_list:
        return {}

    query = """
        SELECT
            employee,
            employee_name,
            shift,
            DATE(`time`) AS attendance_date,
            `time`,
            log_type,
            skip_auto_attendance
        FROM
            `tabEmployee Checkin`
        WHERE
            employee IN %(employee_list)s
            AND DATE(`time`) BETWEEN %(from_date)s AND %(to_date)s
            AND skip_auto_attendance = 0
        ORDER BY
            employee, `time`
    """

    rows = frappe.db.sql(
        query,
        {
            "employee_list": tuple(employee_list),
            "from_date": payrollFrom,
            "to_date": payrollTo,
        },
        as_dict=True,
    )

    # Structure: { employee_id: [punches] }
    attendance_dict = defaultdict(list)
    for row in rows:
        attendance_dict[row["employee"]].append(row)

    return dict(attendance_dict)


# --- Processors ---


import io
import re
from datetime import datetime
from openpyxl import load_workbook
import frappe


TIME_PATTERN = re.compile(r"\d{2}:\d{2}")  # matches 09:55, 19:04 etc


from datetime import datetime, time


def _to_ampm(time_24):
    """
    Convert 24h time to 12h AM/PM.
    Supports:
      - 'HH:MM'
      - 'HH:MM:SS'
      - datetime
      - datetime.time
    Returns None if invalid.
    """

    if not time_24:
        return None

    # If already datetime
    if isinstance(time_24, datetime):
        return time_24.strftime("%I:%M %p")

    # If time object
    if isinstance(time_24, time):
        return datetime.combine(datetime.today(), time_24).strftime("%I:%M %p")

    # String handling
    time_str = str(time_24).strip()

    formats = ("%H:%M", "%H:%M:%S")

    for fmt in formats:
        try:
            return datetime.strptime(time_str, fmt).strftime("%I:%M %p")
        except ValueError:
            continue

    return None


def process_pinnacle(file):
    file_stream = file.stream.read()
    wb = load_workbook(filename=io.BytesIO(file_stream), data_only=True)

    if "Att.log report" not in wb.sheetnames:
        frappe.throw("Sheet 'Att.log report' not found.")

    ws = wb["Att.log report"]

    # ------------------------------------------------
    # DATE LOGIC (COPIED FROM OLD FUNCTION)
    # ------------------------------------------------
    raw_period = ws["C3"].value
    start_date_str = raw_period.split("~")[0].strip()
    start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
    formatted_period = start_date.strftime("%b-%Y")

    # Day numbers (1..31)
    dates = [cell.value for cell in ws[4] if isinstance(cell.value, int)]

    records = []
    max_row = ws.max_row

    row = 5  # first employee row

    while row <= max_row:

        if ws.cell(row=row, column=1).value == "ID:":
            device_id = str(ws.cell(row=row, column=3).value or "").strip()
            emp_name = str(ws.cell(row=row, column=11).value or "").strip()
            punch_row = row + 1

            for col_index, day in enumerate(dates, start=1):
                punch_cell = ws.cell(punch_row, col_index).value

                if not punch_cell:
                    continue

                if not isinstance(punch_cell, str):
                    punch_cell = str(punch_cell)

                times = TIME_PATTERN.findall(punch_cell.replace(" ", ""))
                if not times:
                    continue

                # Build attendance date (✅ copied logic)
                try:
                    attendance_date = datetime.strptime(
                        f"{day:02d}-{formatted_period}", "%d-%b-%Y"
                    ).date()
                except Exception:
                    continue

                for t in times:
                    records.append(
                        {
                            "device_id": device_id,
                            "name": emp_name,
                            "device_name": "Zaicom",
                            "attendance_date": attendance_date,
                            "time": t,
                        }
                    )

            row += 2
        else:
            row += 1

    return records


def process_Opticode_final(file):
    file_stream = file.stream.read()
    wb = load_workbook(filename=io.BytesIO(file_stream), data_only=True)

    if "Final" not in wb.sheetnames:
        frappe.throw("Sheet 'Final' not found in the workbook.")

    sheet = wb["Final"]
    if sheet.max_row < 2:
        return []

    records = []

    # Read header
    headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]
    col = {h: i for i, h in enumerate(headers)}

    required = ["ID", "G", "Date", "In Time", "Out Time"]
    for r in required:
        if r not in col:
            frappe.throw(f"Missing required column: {r}")

    for row in sheet.iter_rows(min_row=2, values_only=True):
        try:
            device_id = row[col["ID"]]
            emp_name = row[col["G"]]
            raw_date = row[col["Date"]]
            in_val = row[col["In Time"]]
            out_val = row[col["Out Time"]]

            if not device_id or not emp_name or not raw_date:
                continue

            # ---- Date ----
            try:
                attendance_date = (
                    raw_date.date()
                    if isinstance(raw_date, datetime)
                    else getdate(raw_date)
                )
            except Exception:
                continue

            # ---- IN punch ----
            in_time = parse_time_safe(in_val)
            if in_time:
                records.append(
                    {
                        "device_id": str(device_id),
                        "name": emp_name.strip(),
                        "device_name": "ESSL Westcott",
                        "attendance_date": attendance_date,
                        "time": in_time.strftime("%H:%M:%S"),
                    }
                )

            # ---- OUT punch ----
            out_time = parse_time_safe(out_val)
            if out_time:
                records.append(
                    {
                        "device_id": str(device_id),
                        "name": emp_name.strip(),
                        "device_name": "ESSL Westcott",
                        "attendance_date": attendance_date,
                        "time": out_time.strftime("%H:%M:%S"),
                    }
                )

        except Exception as e:
            frappe.log_error(
                f"Opticode Final processing error: {e}",
                "Opticode Formatter",
            )

    return records


def process_mantra(file):
    import io
    from datetime import datetime, time
    from openpyxl import load_workbook
    import frappe

    file_stream = file.stream.read()
    wb = load_workbook(filename=io.BytesIO(file_stream), data_only=True)
    sheet = wb.active

    # ---- Header handling ----
    raw_header = [c.value for c in sheet[1]]
    header = merge_header_cells(raw_header)
    header = [str(h).lower().strip() for h in header]

    col_idx = {h: i for i, h in enumerate(header)}

    required = [
        "attendance device id",
        "attendance device",
        "employee name",
        "attendance date",
        "in time",
        "out time",
    ]

    for col in required:
        if col not in col_idx:
            frappe.throw(f"Missing required column: {col}")

    records = []
    seen = set()

    # ---- Helpers ----
    def fmt_date(val):
        if isinstance(val, datetime):
            return val.strftime("%d-%b-%Y")
        try:
            return datetime.strptime(str(val), "%m/%d/%Y").strftime("%d-%b-%Y")
        except Exception:
            return str(val)

    def fmt_time(val):
        if val in (None, "", " "):
            return ""
        if isinstance(val, (datetime, time)):
            return val.strftime("%H:%M")
        return str(val).strip()

    # ---- Row processing ----
    for row in sheet.iter_rows(min_row=2, values_only=True):
        device_id = str(row[col_idx["attendance device id"]] or "").strip()
        device = str(row[col_idx["attendance device"]] or "").strip()
        emp_name = str(row[col_idx["employee name"]] or "").strip()
        date_val = row[col_idx["attendance date"]]

        in_val = row[col_idx["in time"]]
        out_val = row[col_idx["out time"]]

        # Minimum required fields
        if not device_id or not emp_name or not date_val:
            continue

        attendance_date = fmt_date(date_val)
        in_time = fmt_time(in_val)
        out_time = fmt_time(out_val)

        key = (device_id, emp_name, attendance_date, in_time, out_time)
        if key in seen:
            continue
        seen.add(key)

        records.append(
            {
                "device_id": device_id,
                "device": device,
                "employee_name": emp_name,
                "department": "",
                "attendance_date": attendance_date,
                "shift": "",
                "in_time": in_time,
                "out_time": out_time,
            }
        )

    frappe.msgprint(f"✅ Processed {len(records)} Mantra records")
    return records


def process_other(file):
    import io
    from datetime import datetime, time
    from openpyxl import load_workbook
    import frappe

    file_stream = file.stream.read()
    wb = load_workbook(filename=io.BytesIO(file_stream), data_only=True)
    sheet = wb.active

    if sheet.max_row < 2:
        return []

    records = []
    seen = set()

    # ---- Header handling ----
    header_row = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]
    col_index = {h: i for i, h in enumerate(header_row)}

    required_fields = [
        "Employee",
        "Employee Name",
        "Attendance Date",
        "In Time",
        "Out Time",
    ]
    for field in required_fields:
        if field not in col_index:
            frappe.throw(f"Missing required column: {field}")

    # ---- Helpers ----
    def fmt_date(val):
        if isinstance(val, datetime):
            return val.strftime("%d-%b-%Y")
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(str(val), fmt).strftime("%d-%b-%Y")
            except Exception:
                pass
        return str(val)

    def fmt_time(val):
        if val in (None, "", " "):
            return ""
        if isinstance(val, (datetime, time)):
            return val.strftime("%H:%M")
        return str(val).strip()

    # ---- Row processing ----
    for row in sheet.iter_rows(min_row=2, values_only=True):
        try:
            employee = row[col_index["Employee"]]
            emp_name = row[col_index["Employee Name"]]
            date_val = row[col_index["Attendance Date"]]
            in_val = row[col_index["In Time"]]
            out_val = row[col_index["Out Time"]]

            if not employee or not emp_name or not date_val:
                continue

            attendance_date = fmt_date(date_val)
            in_time = fmt_time(in_val)
            out_time = fmt_time(out_val)

            key = (str(employee), attendance_date, in_time, out_time)
            if key in seen:
                continue
            seen.add(key)

            records.append(
                {
                    "employee_id": str(employee),
                    "employee_name": emp_name.strip(),
                    "attendance_date": attendance_date,
                    "shift": "Regular",
                    "in_time": in_time,
                    "out_time": out_time,
                }
            )

        except Exception as e:
            frappe.log_error(
                f"Error processing Other attendance row: {e}", "Attendance Formatter"
            )

    return records


# --- Core Final Generator ---
@frappe.whitelist()
def generate_final_sheet(attendance_data=None):
    """
    Generate final punch-style attendance from mixed raw attendance inputs.

    Supports:
    - Raw punch logs (time)
    - IN/OUT based logs (in_time, out_time)

    Output:
        Employee | Name | Date | Shift | Log Type | Time | Punch From
    """

    attendance_data = attendance_data or []

    # employee -> date -> punches
    punch_map = defaultdict(lambda: defaultdict(list))

    for row in attendance_data:
        try:
            # -------------------------------
            # Normalize employee & device
            # -------------------------------
            emp_name = row.get("name") or row.get("employee_name")
            device = row.get("device_name") or row.get("device") or "N/A"
            device_id = row.get("device_id")

            if not emp_name:
                continue

            # -------------------------------
            # Normalize date
            # -------------------------------
            try:
                punch_date = getdate(row.get("attendance_date"))
            except Exception:
                continue

            # --------------------------------
            # CASE 1: Raw punch-style data
            # --------------------------------
            if row.get("time"):
                punch_time = parse_time_safe(row.get("time"))
                if not punch_time:
                    continue

                punch_map[emp_name][punch_date].append(
                    {
                        "time": punch_time,
                        "device": device,
                        "device_id": device_id,
                    }
                )

            # --------------------------------
            # CASE 2: IN / OUT based data
            # --------------------------------
            else:
                in_time = parse_time_safe(row.get("in_time"))
                out_time = parse_time_safe(row.get("out_time"))

                if in_time:
                    punch_map[emp_name][punch_date].append(
                        {
                            "time": in_time,
                            "device": device,
                            "device_id": device_id,
                        }
                    )

                if out_time:
                    punch_map[emp_name][punch_date].append(
                        {
                            "time": out_time,
                            "device": device,
                            "device_id": device_id,
                        }
                    )

        except Exception:
            frappe.log_error(frappe.get_traceback(), f"Punch parsing error: {row}")

    # ----------------------------------
    # BUILD FINAL RESULT
    # ----------------------------------
    final_rows = []

    for emp_name, dates in punch_map.items():
        for date_key, punches in dates.items():

            if not punches:
                continue

            # Sort punches by time
            punches.sort(key=lambda x: x["time"])

            first_punch = punches[0]
            last_punch = punches[-1]
            # -------- IN --------
            final_rows.append(
                {
                    "employee": _get_emp_id(
                        first_punch["device"], first_punch.get("device_id")
                    ),
                    "employee_name": emp_name,
                    "attendance_date": date_key,
                    "shift": frappe.get_value(
                        "Employee",
                        _get_emp_id(
                            first_punch["device"], first_punch.get("device_id")
                        ),
                        "default_shift",
                    )
                    or "Regular",
                    "log_type": "IN",
                    "time": first_punch["time"].strftime("%H:%M:%S"),
                    "punch_from": first_punch["device"],
                }
            )

            # -------- OUT (only if different) --------
            if first_punch["time"] != last_punch["time"]:
                final_rows.append(
                    {
                        "employee": _get_emp_id(
                            last_punch["device"], last_punch.get("device_id")
                        ),
                        "employee_name": emp_name,
                        "attendance_date": date_key,
                        "shift": frappe.get_value(
                            "Employee",
                            _get_emp_id(
                                first_punch["device"], first_punch.get("device_id")
                            ),
                            "default_shift",
                        )
                        or "Regular",
                        "log_type": "OUT",
                        "time": last_punch["time"].strftime("%H:%M:%S"),
                        "punch_from": last_punch["device"],
                    }
                )

    return {
        "message": "✅ Final attendance generated",
        "total_records": len(final_rows),
        "data": final_rows,
    }


# --- Load Raw Data --- #
@frappe.whitelist()
def load_raw_attendance_data():
    pinnacle_file = frappe.request.files.get("pinnacle_file")
    opticode_file = frappe.request.files.get("opticode_file")
    mantra_file = frappe.request.files.get("mantra_file")
    other_file = frappe.request.files.get("other_file")

    # company = frappe.form_dict.get("company") or None
    payrollFrom = frappe.form_dict.get("from_date")
    payrollTo = frappe.form_dict.get("to_date")

    if not payrollFrom or not payrollTo:
        return Response("❌ Company and Payroll Period are required", status=400)

    # Fetch employees for the selected company
    employeeList = get_employee()

    if not employeeList:
        return Response("❌ No active employees found for this company", status=404)

    # Fetch attendance from the app for given employees and period
    app_attendance = get_app_attendance(
        list(employeeList.keys()), payrollFrom, payrollTo
    )

    # Convert app_attendance to same structure as other records
    app_records = convert_app_attendance_to_records(app_attendance)

    if pinnacle_file:
        pinnacleAttendance = process_pinnacle(pinnacle_file)
    if opticode_file:
        opticodeAttendance = process_Opticode_final(opticode_file)
    if mantra_file:
        mantraAttendance = process_mantra(mantra_file)
    if other_file:
        otherFile = process_other(other_file)
    appAttendance = app_records

    return {
        "message": "✅ Attendance files processed successfully",
        "status_cd": 200,
        "pinnacle_attendance": pinnacleAttendance if pinnacle_file else [],
        "opticode_attendance": opticodeAttendance if opticode_file else [],
        "mantra_attendance": mantraAttendance if mantra_file else [],
        "other_attendance": otherFile if other_file else [],
        "app_attendance": appAttendance,
    }


# --- Preview and Download Endpoints ---
@frappe.whitelist()
def preview_final_attendance_sheet(raw_data=None):
    """
    Preview final attendance in punch-style rows:
    Employee | Name | Date | Shift | Log Type | Time | Punch From
    """

    if not raw_data:
        return {
            "message": "No raw data provided",
            "data": [],
            "total_records": 0,
        }

    # raw_data may come as JSON string OR dict
    if isinstance(raw_data, str):
        raw_data = json.loads(raw_data)

    # Collect all raw records
    records = []
    records += raw_data.get("pinnacle", [])
    records += raw_data.get("opticode", [])
    records += raw_data.get("mantra", [])
    records += raw_data.get("other", [])
    records += raw_data.get("app", [])

    # Generate final punch-style rows (LIST)
    result = generate_final_sheet(records)
    # frappe.throw(str(result))
    final_rows = result.get("data", [])
    return {
        "message": "Preview generated successfully",
        "total_records": len(final_rows),
        "data": final_rows,
    }


@frappe.whitelist()
def download_final_attendance_excel(logs):
    import json, io
    from datetime import datetime
    from openpyxl import Workbook
    from werkzeug.wrappers import Response

    raw_data = json.loads(logs)

    # -----------------------------
    # 1. NORMALIZE DATA
    # -----------------------------
    rows = []

    # Case A: dict grouped by employee
    if isinstance(raw_data, dict):
        for emp_id, emp_rows in raw_data.items():
            for row in emp_rows:
                rows.append(row)

    # Case B: flat list
    elif isinstance(raw_data, list):
        rows = raw_data

    else:
        frappe.throw("Unsupported attendance data format")

    if not rows:
        frappe.throw("No attendance data found")

    # -----------------------------
    # 2. SAFE SORTING
    # -----------------------------
    rows = sorted(
        rows,
        key=lambda x: (
            x.get("employee") or x.get("employee_name", ""),
            x.get("attendance_date", ""),
            x.get("time", ""),
        ),
    )

    # -----------------------------
    # 3. CREATE EXCEL
    # -----------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "Final Attendance"

    ws.append(
        [
            "Employee",
            "Employee Name",
            "Attendance Date",
            "Shift",
            "Log Type",
            "Time",
            "Punch From",
        ]
    )

    for row in rows:
        attendance_date = row.get("attendance_date")

        if isinstance(attendance_date, datetime):
            attendance_date = attendance_date.strftime("%Y-%m-%d")
        else:
            attendance_date = str(attendance_date or "")

        ws.append(
            [
                row.get("employee", ""),  # may be empty
                row.get("employee_name", ""),
                attendance_date,
                row.get("shift", ""),
                row.get("log_type", ""),
                row.get("time", ""),
                row.get("punch_from", ""),
            ]
        )

    # -----------------------------
    # 4. RETURN FILE
    # -----------------------------
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Final_Attendance.xlsx"},
    )


@frappe.whitelist()
def create_data_import_for_attendance(attendance_data=None):
    """
    Creates a Data Import record with an Excel file for Attendance List
    and starts the import process.
    """
    # 1. Fetch validated records
    validated_data = json.loads(attendance_data or "[]")
    if not validated_data:
        frappe.throw("No validated records found.")

    # 2. Create Excel file in memory
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance List"

    headers = [
        "Employee",
        "Employee Name",
        "Shift",
        "Log Type",
        "Time",
        "Punch From",
    ]
    ws.append(headers)

    for emp_id in sorted(validated_data):
        for row in validated_data[emp_id]:
            ws.append(
                [
                    row["employee"],
                    row["employee_name"],
                    row.get("shift"),
                    row.get("log_type"),
                    f"{row.get('attendance_date', '')} {row.get('time', '')}",
                    row.get("punch_from"),
                ]
            )

    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # 3. Create Data Import record first
    data_import = frappe.get_doc(
        {
            "doctype": "Data Import",
            "import_type": "Insert New Records",
            "reference_doctype": "Employee Checkin",
            "submit_after_import": 0,
            "mute_emails": 1,
        }
    )
    data_import.insert(ignore_permissions=True)

    # 4. Attach the Excel file to Data Import
    file_doc = save_file(
        "attendance_import.xlsx",
        output.getvalue(),
        "Data Import",
        data_import.name,
        is_private=1,
    )
    data_import.import_file = file_doc.file_url
    data_import.save(ignore_permissions=True)

    # 5. Start the import
    frappe.enqueue(
        "frappe.core.doctype.data_import.data_import.start_import",
        data_import=data_import.name,
        import_type="Insert New Records",
    )

    return data_import.name


# ============================================================
# ------------------ Validated Records -----------------------
# ============================================================


@frappe.whitelist()
def validate_attendance_data(attendance_data=None):

    data = _coerce_json_arg(attendance_data)
    # print(data)
    if not data:
        return {"validated": {}, "non_validated": []}

    # Normalize input
    if isinstance(data, dict):
        rows = []
        for emp in data:
            rows.extend(data[emp])
    else:
        rows = data

    validated = defaultdict(list)
    non_validated = []
    seen_dates = set()

    # ---------------- RULES ----------------

    def is_empty_or_zero(t):
        if t is None:
            return True
        t = str(t).strip()
        return t == "" or t in ("0", "00", "00:00", "00:00:00")

    def rule_missing_punch(r):

        if is_empty_or_zero(r.get("time")) or is_empty_or_zero(r.get("time")):
            return False, "Missing time"
        return True, None

    def rule_missing_employee(r):
        if not r.get("employee") or not r.get("employee_name"):
            return False, "Missing employee information"
        return True, None

    def rule_invalid_time(r):
        try:
            time_str = r.get("time")

            if is_empty_or_zero(time_str):
                return True, None

            time_obj = parse_time_safe(str(time_str))
            if time_obj is None:
                return False, "Invalid time format"

        except Exception:
            return False, "Invalid time format"

        return True, None

    def rule_duplicate(r):
        key = f"{r.get('employee')}_{r.get('attendance_date')}_{r.get('time')}"
        if key in seen_dates:
            return False, "Duplicate entry"
        seen_dates.add(key)
        return True, None

    # ---------------- PIPELINE ----------------

    validations = [
        rule_missing_punch,
        rule_invalid_time,
        rule_duplicate,
        rule_missing_employee,
    ]

    for r in rows:
        errors = []

        for check in validations:
            ok, msg = check(r)
            if not ok:
                errors.append(msg)

        if errors:
            r["errors"] = errors
            non_validated.append(r)
        else:
            validated[r.get("employee")].append(r)

    return {
        "message": "Validation completed",
        "validated": validated,
        "non_validated": non_validated,
        "total_valid": sum(len(v) for v in validated.values()),
        "total_invalid": len(non_validated),
    }
